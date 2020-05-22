import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import datetime
from pathlib import Path
from audit_inspector.common import settings
from collections import Counter
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint

def main(results, evidence_dir):
    wb = openpyxl.Workbook() # Initalize an empty workbook
    create_platform_sheets(wb, results)
    write_results_data(wb, results)
    wb.save(evidence_dir / 'Test Results.xlsx')


def create_platform_sheets(wb, results):
    #write_skeleton(wb, results):
    """
    1. Iterate through the results dictionary and create one sheet per platform.
    2. On each platform sheet, write the dict keys as a header row.
    2. Write evidence metadata on the overview tab that provides high-level status and issues.
    """
    sheet = wb.active
    sheet.title = 'Overview'
    today_raw = datetime.date.today()
    evidence_dates = set()
    
    # Iterate through the results and create one sheet per platform.
    for dict_item in results:
        control = dict_item['Control']
        if dict_item['Platform'] not in wb.sheetnames:
            ws = wb.create_sheet(dict_item['Platform'])
        if dict_item['Date'] not in evidence_dates:
            evidence_dates.add(dict_item['Date'])

    # Write evidence metadata on the overview tab that provides high-level status and issues.
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
    sheet['A1'].fill = settings.white_fill
    sheet['A1'].alignment = Alignment(wrapText=True)
    sheet['A1'].value = f"Report Run Date: {today_raw.strftime('%m/%d/%Y')}\nEvidence Dates: {', '.join(sorted(evidence_dates))}"

    # Write header rows on each platform tab        
    column_headers = set_column_headers(control)
    longest = len(max(column_headers, key=len)) + 10
    
    for sheet in wb.worksheets:
        column_num = 1
        if 'Overview' not in sheet.title:
            for col, val in enumerate(column_headers, start=1):
                column_num = report_write_headers(sheet, column_num, val, longest)


def write_results_data(wb, results):
    for p in settings.platforms:
        row_num = 2
        for sheet in wb.worksheets:
            if sheet.title == p:
                for dict_item in results:
                    if dict_item['Platform'] == p:
                        for k,v in dict_item.items():
                            for row in sheet.iter_rows(min_col=1, min_row=row_num, max_col=sheet.max_column, max_row=row_num):
                                for cell in row:
                                    if sheet.cell(1, cell.column).value == k: # If the Key matches the column header
                                        if isinstance(v, list):
                                            cell.value = '\n'.join(v)
                                        else:
                                            cell.value = str(v)
                                        if k == 'Notes':
                                            cell.value = '\n'.join(v).split('$$')[1]
                                    cell.alignment = Alignment(wrapText=True)
                                    cell.font = Font(size=9)
                        row_num += 1


def report_write_headers(sheet, column_num, heading, longest):
    sheet.cell(row=1, column=column_num).value = heading
    sheet.cell(row=1, column=column_num).font = settings.header_font
    sheet.cell(row=1, column=column_num).fill = settings.dark_blue_fill
    column_letter = get_column_letter(column_num)
    sheet.column_dimensions[column_letter].width = longest
    column_num += 1
    return column_num


def set_column_headers(control):
    """
    based on the control type
    """
    column_headers = []
    for key in settings.report_headers.keys():
        if control == key:
            column_headers = settings.report_headers[key]
    return column_headers
