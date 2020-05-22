import openpyxl
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

control_categories = {
    'kubernetes': {
        'connection': 'openssl s_client',
        'firewall': 'get networkpolicy',
        'patching': '+ kubectl version'
    },
    'linux': {
        'connection': 'MSG_SERVICE_ACCEPT',
        'firewall': 'iptables -l'
    }
}

insecure_ciphers = {
    'ssh':
        [
        'arcfour',
        'arcfour256',
        'arcfour128'
        ],
    'tls':
        [
        'TLSv1'
        ]
    }

report_headers = {
    'connection': ['Hostname', 'Protocol', 'Version', 'Cipher', 'Available Ciphers', 'Credential Methods', 'Idle Timeout', 'Root Login', 'Notes']
}

platforms = ['Kubernetes', 'Linux']


# Excel formatting variables
make_bold = Font(bold=True)
make_italic = Font(italic=True)
dark_green_fill = PatternFill(
    start_color='03990f', end_color='03990f', fill_type='solid')
light_gray_fill = PatternFill(
    start_color='bcb7b9', end_color='bcb7b9', fill_type='solid')
light_purple_fill = PatternFill(
    start_color='d2a2f2', end_color='d2a2f2', fill_type='solid')
light_green_fill = PatternFill(
    start_color='8af202', end_color='8af202', fill_type='solid')
light_pink_fill = PatternFill(
    start_color='f7b2f2', end_color='f7b2f2', fill_type='solid')
light_blue_fill = PatternFill(
    start_color='02f2ee', end_color='02f2ee', fill_type='solid')
light_orange_fill = PatternFill(
    start_color='efc25f', end_color='efc25f', fill_type='solid')
light_yellow_fill = PatternFill(
    start_color='eff24d', end_color='eff24d', fill_type='solid')
lime_green_fill = PatternFill(
    start_color='ccffcc', end_color='ccffcc', fill_type='solid')
bright_pink_fill = PatternFill(
    start_color='ffccff', end_color='ffccff', fill_type='solid')
bright_red_fill = PatternFill(
    start_color='ff0000', end_color='ff0000', fill_type='solid')
white_fill = PatternFill(
    start_color='ffffff', end_color='ffffff', fill_type='solid')
dark_blue_fill = PatternFill(
    start_color='000066', end_color='000066', fill_type='solid')
no_fill = openpyxl.styles.PatternFill(fill_type=None)
header_font = Font(size=14, underline='single', color='ffffff', bold=True)